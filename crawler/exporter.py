"""数据导出模块 — 多格式导出、数据库导出、大数据量分批处理"""

import csv
import json
import logging
import os
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class DataExporter:
    """
    数据导出器

    支持功能:
    - 多格式导出：CSV、JSON、JSONL、Excel、SQLite、Markdown
    - 数据库导出：MySQL（可选）、SQLite、MongoDB（可选）
    - 大数据量分批导出（避免内存溢出）
    - 自动字段映射与重命名
    - 导出模板（预定义字段子集）
    - 导出统计与进度追踪
    """

    def __init__(self, config: dict):
        self.config = config
        exporter_cfg = config.get("exporter", {})

        # 默认配置
        self._batch_size: int = exporter_cfg.get("batch_size", 1000)
        self._default_format: str = exporter_cfg.get("default_format", "json")
        self._encoding: str = exporter_cfg.get("encoding", "utf-8")
        self._ensure_ascii: bool = exporter_cfg.get("ensure_ascii", False)

        # 统计
        self._stats: Dict[str, int] = {
            "total_exported": 0,
            "batches": 0,
            "errors": 0,
        }

        logger.info("📤 DataExporter 初始化完成")

    # ──────────────────────────────
    # 通用导出入口
    # ──────────────────────────────

    def export(
        self,
        data: List[Dict],
        output_path: str,
        fmt: Optional[str] = None,
        fields: Optional[List[str]] = None,
        field_mapping: Optional[Dict[str, str]] = None,
        batch_size: Optional[int] = None,
    ) -> Dict[str, Any]:
        """通用导出接口

        Args:
            data: 数据列表
            output_path: 输出路径
            fmt: 导出格式（None 则根据扩展名推断）
            fields: 导出字段列表（None 则导出全部）
            field_mapping: 字段名映射 {原字段: 新字段}
            batch_size: 分批大小

        Returns:
            导出统计信息
        """
        # 推断格式
        if fmt is None:
            fmt = self._infer_format(output_path)

        # 字段处理
        if fields or field_mapping:
            data = self._process_fields(data, fields, field_mapping)

        # 分批导出
        batch_size = batch_size or self._batch_size

        exporters = {
            "json": self._export_json,
            "jsonl": self._export_jsonl,
            "csv": self._export_csv,
            "sqlite": self._export_sqlite,
            "markdown": self._export_markdown,
            "excel": self._export_excel,
            "mysql": self._export_mysql,
            "mongodb": self._export_mongodb,
        }

        exporter_func = exporters.get(fmt)
        if not exporter_func:
            raise ValueError(f"不支持的导出格式: {fmt}，支持: {', '.join(exporters.keys())}")

        # 确保输出目录存在
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"📤 开始导出: {len(data)} 条数据 → {output_path} (格式: {fmt})")

        try:
            if len(data) > batch_size and fmt in ("jsonl", "csv", "sqlite"):
                # 大数据分批写入
                result = exporter_func(data, output_path, batch_size=batch_size)
            else:
                result = exporter_func(data, output_path)

            self._stats["total_exported"] += len(data)
            logger.info(f"✅ 导出完成: {len(data)} 条数据 → {output_path}")
            return result

        except Exception as e:
            self._stats["errors"] += 1
            logger.error(f"导出失败: {e}")
            raise

    # ──────────────────────────────
    # JSON 导出
    # ──────────────────────────────

    def _export_json(self, data: List[Dict], output_path: str, **kwargs) -> Dict:
        """导出为 JSON 文件"""
        with open(output_path, "w", encoding=self._encoding) as f:
            json.dump(data, f, ensure_ascii=self._ensure_ascii, indent=2, default=str)
        self._stats["batches"] += 1
        return {"format": "json", "count": len(data), "path": output_path}

    # ──────────────────────────────
    # JSONL 导出（逐行 JSON，适合大数据量）
    # ──────────────────────────────

    def _export_jsonl(
        self,
        data: List[Dict],
        output_path: str,
        batch_size: int = 1000,
        **kwargs,
    ) -> Dict:
        """导出为 JSONL 文件（每行一个 JSON 对象）"""
        total = 0
        with open(output_path, "w", encoding=self._encoding) as f:
            for i in range(0, len(data), batch_size):
                batch = data[i:i + batch_size]
                for item in batch:
                    f.write(json.dumps(item, ensure_ascii=self._ensure_ascii, default=str) + "\n")
                total += len(batch)
                self._stats["batches"] += 1
                logger.debug(f"JSONL 已写入 {total}/{len(data)} 条")
        return {"format": "jsonl", "count": total, "path": output_path}

    # ──────────────────────────────
    # CSV 导出
    # ──────────────────────────────

    def _export_csv(
        self,
        data: List[Dict],
        output_path: str,
        batch_size: int = 1000,
        **kwargs,
    ) -> Dict:
        """导出为 CSV 文件"""
        if not data:
            return {"format": "csv", "count": 0, "path": output_path}

        # 收集所有字段名
        all_keys = []
        seen = set()
        for item in data:
            for key in item.keys():
                if key not in seen:
                    all_keys.append(key)
                    seen.add(key)

        csv_cfg = self.config.get("storage", {}).get("csv", {})
        delimiter = csv_cfg.get("delimiter", ",")

        total = 0
        with open(output_path, "w", encoding=self._encoding, newline="") as f:
            writer = csv.DictWriter(f, fieldnames=all_keys, delimiter=delimiter, extrasaction="ignore")
            writer.writeheader()

            for i in range(0, len(data), batch_size):
                batch = data[i:i + batch_size]
                for item in batch:
                    # 将复杂类型转为字符串
                    row = {}
                    for k, v in item.items():
                        if isinstance(v, (list, dict)):
                            row[k] = json.dumps(v, ensure_ascii=False)
                        else:
                            row[k] = v
                    writer.writerow(row)
                total += len(batch)
                self._stats["batches"] += 1

        return {"format": "csv", "count": total, "path": output_path, "fields": all_keys}

    # ──────────────────────────────
    # SQLite 导出
    # ──────────────────────────────

    def _export_sqlite(
        self,
        data: List[Dict],
        output_path: str,
        batch_size: int = 1000,
        **kwargs,
    ) -> Dict:
        """导出为 SQLite 数据库"""
        if not data:
            return {"format": "sqlite", "count": 0, "path": output_path}

        sqlite_cfg = self.config.get("storage", {}).get("sqlite", {})
        table_name = sqlite_cfg.get("table_name", "crawled_data")

        # 收集所有字段
        all_keys = []
        seen = set()
        for item in data:
            for key in item.keys():
                if key not in seen:
                    all_keys.append(key)
                    seen.add(key)

        conn = sqlite3.connect(output_path)
        cursor = conn.cursor()

        try:
            # 创建表
            columns = ", ".join(f'"{k}" TEXT' for k in all_keys)
            cursor.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({columns})')

            # 分批插入
            placeholders = ", ".join(["?"] * len(all_keys))
            col_names = ", ".join(f'"{k}"' for k in all_keys)
            insert_sql = f'INSERT INTO "{table_name}" ({col_names}) VALUES ({placeholders})'

            total = 0
            for i in range(0, len(data), batch_size):
                batch = data[i:i + batch_size]
                rows = []
                for item in batch:
                    row = []
                    for key in all_keys:
                        value = item.get(key)
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        elif value is not None:
                            value = str(value)
                        row.append(value)
                    rows.append(row)

                cursor.executemany(insert_sql, rows)
                total += len(rows)
                self._stats["batches"] += 1

            conn.commit()
            logger.info(f"SQLite 已写入 {total} 条到 {table_name}")

        finally:
            conn.close()

        return {"format": "sqlite", "count": total, "path": output_path, "table": table_name}

    # ──────────────────────────────
    # Markdown 导出
    # ──────────────────────────────

    def _export_markdown(self, data: List[Dict], output_path: str, **kwargs) -> Dict:
        """导出为 Markdown 文件"""
        lines = []
        lines.append(f"# 爬取数据导出\n")
        lines.append(f"> 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        lines.append(f"> 数据量: {len(data)} 条\n\n")
        lines.append("---\n\n")

        for i, item in enumerate(data, 1):
            lines.append(f"## 条目 {i}\n\n")
            for key, value in item.items():
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False, indent=2)
                lines.append(f"**{key}**: {value}\n\n")
            lines.append("---\n\n")

        with open(output_path, "w", encoding=self._encoding) as f:
            f.write("".join(lines))

        self._stats["batches"] += 1
        return {"format": "markdown", "count": len(data), "path": output_path}

    # ──────────────────────────────
    # Excel 导出（可选依赖 openpyxl）
    # ──────────────────────────────

    def _export_excel(self, data: List[Dict], output_path: str, **kwargs) -> Dict:
        """导出为 Excel 文件（需要 openpyxl）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.warning("openpyxl 未安装，降级为 CSV 导出")
            csv_path = output_path.rsplit(".", 1)[0] + ".csv"
            return self._export_csv(data, csv_path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "爬取数据"

        if not data:
            wb.save(output_path)
            return {"format": "excel", "count": 0, "path": output_path}

        # 收集字段
        all_keys = []
        seen = set()
        for item in data:
            for key in item.keys():
                if key not in seen:
                    all_keys.append(key)
                    seen.add(key)

        # 写表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, key in enumerate(all_keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 写数据
        for row_idx, item in enumerate(data, 2):
            for col_idx, key in enumerate(all_keys, 1):
                value = item.get(key)
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(output_path)
        self._stats["batches"] += 1
        return {"format": "excel", "count": len(data), "path": output_path}

    # ──────────────────────────────
    # MySQL 导出（可选依赖）
    # ──────────────────────────────

    def _export_mysql(self, data: List[Dict], output_path: str, **kwargs) -> Dict:
        """导出到 MySQL 数据库（需要 pymysql）"""
        try:
            import pymysql
        except ImportError:
            logger.error("pymysql 未安装，无法导出到 MySQL。请安装: pip install pymysql")
            raise ImportError("pymysql 未安装，请运行: pip install pymysql")

        mysql_cfg = self.config.get("exporter", {}).get("mysql", {})
        host = mysql_cfg.get("host", "localhost")
        port = mysql_cfg.get("port", 3306)
        user = mysql_cfg.get("user", "root")
        password = mysql_cfg.get("password", "")
        database = mysql_cfg.get("database", "crawler")
        table_name = mysql_cfg.get("table_name", "crawled_data")
        charset = mysql_cfg.get("charset", "utf8mb4")

        if not data:
            return {"format": "mysql", "count": 0, "table": table_name}

        # 收集字段
        all_keys = list(dict.fromkeys(key for item in data for key in item.keys()))

        conn = pymysql.connect(
            host=host, port=port, user=user,
            password=password, database=database,
            charset=charset,
        )
        cursor = conn.cursor()

        try:
            # 创建表
            columns = ", ".join(f'`{k}` TEXT' for k in all_keys)
            cursor.execute(f'CREATE TABLE IF NOT EXISTS `{table_name}` ({columns}) ENGINE=InnoDB DEFAULT CHARSET={charset}')

            # 分批插入
            placeholders = ", ".join(["%s"] * len(all_keys))
            insert_sql = f'INSERT INTO `{table_name}` ({", ".join(f"`{k}`" for k in all_keys)}) VALUES ({placeholders})'

            total = 0
            for i in range(0, len(data), self._batch_size):
                batch = data[i:i + self._batch_size]
                rows = []
                for item in batch:
                    row = []
                    for key in all_keys:
                        value = item.get(key)
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        elif value is not None:
                            value = str(value)
                        row.append(value)
                    rows.append(row)
                cursor.executemany(insert_sql, rows)
                total += len(rows)
                self._stats["batches"] += 1

            conn.commit()
            logger.info(f"MySQL 已写入 {total} 条到 {table_name}")

        finally:
            conn.close()

        return {"format": "mysql", "count": total, "table": table_name, "database": database}

    # ──────────────────────────────
    # MongoDB 导出（可选依赖）
    # ──────────────────────────────

    def _export_mongodb(self, data: List[Dict], output_path: str, **kwargs) -> Dict:
        """导出到 MongoDB（需要 pymongo）"""
        try:
            import pymongo
        except ImportError:
            logger.error("pymongo 未安装，无法导出到 MongoDB。请安装: pip install pymongo")
            raise ImportError("pymongo 未安装，请运行: pip install pymongo")

        mongo_cfg = self.config.get("exporter", {}).get("mongodb", {})
        uri = mongo_cfg.get("uri", "mongodb://localhost:27017")
        database = mongo_cfg.get("database", "crawler")
        collection_name = mongo_cfg.get("collection", "crawled_data")

        client = pymongo.MongoClient(uri)
        db = client[database]
        collection = db[collection_name]

        try:
            total = 0
            for i in range(0, len(data), self._batch_size):
                batch = data[i:i + self._batch_size]
                # 转换复杂类型
                clean_batch = []
                for item in batch:
                    clean_item = {}
                    for k, v in item.items():
                        if isinstance(v, datetime):
                            clean_item[k] = v.isoformat()
                        else:
                            clean_item[k] = v
                    clean_batch.append(clean_item)

                collection.insert_many(clean_batch)
                total += len(clean_batch)
                self._stats["batches"] += 1

            logger.info(f"MongoDB 已写入 {total} 条到 {database}.{collection_name}")

        finally:
            client.close()

        return {"format": "mongodb", "count": total, "database": database, "collection": collection_name}

    # ──────────────────────────────
    # 工具方法
    # ──────────────────────────────

    @staticmethod
    def _infer_format(path: str) -> str:
        """根据文件扩展名推断格式"""
        ext_map = {
            ".json": "json",
            ".jsonl": "jsonl",
            ".ndjson": "jsonl",
            ".csv": "csv",
            ".tsv": "csv",
            ".db": "sqlite",
            ".sqlite": "sqlite",
            ".sqlite3": "sqlite",
            ".md": "markdown",
            ".markdown": "markdown",
            ".xlsx": "excel",
            ".xls": "excel",
        }
        ext = Path(path).suffix.lower()
        return ext_map.get(ext, "json")

    @staticmethod
    def _process_fields(
        data: List[Dict],
        fields: Optional[List[str]] = None,
        field_mapping: Optional[Dict[str, str]] = None,
    ) -> List[Dict]:
        """字段过滤与映射"""
        result = []
        for item in data:
            new_item = {}

            if fields:
                for field in fields:
                    if field in item:
                        new_item[field] = item[field]
            else:
                new_item = dict(item)

            if field_mapping:
                for old_name, new_name in field_mapping.items():
                    if old_name in new_item:
                        new_item[new_name] = new_item.pop(old_name)

            result.append(new_item)

        return result

    def get_stats(self) -> Dict[str, int]:
        """获取导出统计"""
        return dict(self._stats)
